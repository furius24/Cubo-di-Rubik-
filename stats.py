from typing import Type, Callable
import matplotlib.pyplot as plt

from time import time
from openpyxl import Workbook, load_workbook


from ..cubo_rubik.cubo_storico import CuboStorico
from ..cubo_rubik.risolutore import genera_soluzione
from ..mischia.generatore import gen_seq_mov_casuali
from ..mischia.controllore import trasf_in_movimenti, trasf_in_stringa
from ..mischia.ripulitore import pulisci_movimenti

# Classe utile alle statistiche riguardanti la risoluzione del cubo di Rubik secondo l'algoritmo eseguito
class Statistiche:
    # Definizione del metodo costruttore e passaggio delle singole variabili agli attributi
    def __init__(self, cubo: Type[CuboStorico], mischiare: Callable, risolutore: Callable):
        self.cubo_classe = cubo
        self.mischiare = mischiare
        self.risolutore = risolutore
        # Inizializzazione foglio di lavoro
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    # Generazione titoli identificativi
    def gen_titoli(self):
        titoli = ["Tempo per generare la soluzione", "Mov_per_mischiare", "Lunghezza della soluzione", "Soluzione"]
        if not any(row[0].value == titoli[0] for row in self.worksheet.iter_rows(min_row=1, max_row=1)):
            self.worksheet.append(titoli)

    # Genera statistiche utilizzando i metodi nelle classi esterne
    def genera_statistiche(self, risoluzioni: int):
        # Ottieni l'indice dell'ultima riga con dati nel foglio di lavoro
        last_row = self.worksheet.max_row
        # Ciclo sul numero di risoluzioni richieste
        for _ in range(risoluzioni):
            # Genero il cubo di Rubik
            cubo_corrente = self.cubo_classe(3)
            # Genero la sequenza per mischiare il cubo e la eseguo
            mov_per_mischiare = self.mischiare()
            cubo_corrente.esegui_movimenti(mov_per_mischiare, salva_storico=False)
            # Inizio a contare il tempo
            start_time = time()
            # Genero la soluzione
            Soluzione = trasf_in_movimenti(pulisci_movimenti(
                                        trasf_in_stringa(self.risolutore(cubo_corrente))))
            # Stoppo il tempo
            end_time = time()
            # Calcolo il delta di tempo risultato
            tempo_soluzione = end_time - start_time
            # Aggiungi i nuovi valori all'ultima riga del foglio di lavoro
            self.worksheet.cell(row=last_row + 1, column=1, value=tempo_soluzione)
            self.worksheet.cell(row=last_row + 1, column=2, value=mov_per_mischiare)
            self.worksheet.cell(row=last_row + 1, column=3, value=len(Soluzione))
            self.worksheet.cell(row=last_row + 1, column=4, value=trasf_in_stringa(Soluzione))
            last_row += 1
        # Salva il foglio di lavoro
        self.workbook.save("statistiche_cubo.xlsx")
    
    # Metodo per generare il grafico di dispersione
    def genera_grafico_dispersione(self, file_path):
        # Carica i dati dal foglio di lavoro
        workbook = load_workbook(file_path)
        worksheet = workbook.active
        # Estrai i valori dalla prima e dalla terza colonna
        colonna_tempo = [cella.value for cella in worksheet["A"][1:]]
        colonna_lunghezza = [cella.value for cella in worksheet["C"][1:]]
        # Crea il grafico di dispersione
        plt.figure(figsize=(8, 6))  # Dimensione del grafico
        plt.scatter(colonna_tempo, colonna_lunghezza, marker="o", color="blue", s=50, label="Dati")  # Personalizzazioni dei punti
        plt.xlabel("Tempo per generare la soluzione (secondi)", fontdict={"fontsize": 12})  # Etichetta dell'asse X
        plt.ylabel("Lunghezza della soluzione (mosse)", fontdict={"fontsize": 12})  # Etichetta dell'asse Y
        plt.title("Grafico di dispersione", fontdict={"fontsize": 14})  # Titolo del grafico
        plt.grid(True)  # Griglia di sfondo
        plt.legend()  # Legenda
        plt.xticks(fontsize=10)  # Personalizzazioni degli assi X
        plt.yticks(fontsize=10)  # Personalizzazioni degli assi Y
        plt.show()  # Mostra il grafico
    
    # Metodo per aprire il file del foglio di lavoro
    def apri_workbook(self, percorso_file):
        self.workbook = Workbook()
        self.workbook = load_workbook(percorso_file)
        self.worksheet = self.workbook.active

# Test statistiche
if __name__ == "__main__":
    stats = Statistiche(CuboStorico, gen_seq_mov_casuali, genera_soluzione)
    stats.apri_workbook("statistiche_cubo.xlsx")
    stats.gen_titoli()
    stats.genera_statistiche(200)
    stats.genera_grafico_dispersione("statistiche_cubo.xlsx")